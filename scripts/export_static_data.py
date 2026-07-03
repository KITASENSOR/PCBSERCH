import argparse
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def read_rows(ws, columns):
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [cell_text(row[i]) if i < len(row) else "" for i in range(columns)]
        if any(values):
            yield values


def unique_list(values):
    seen = set()
    result = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def split_parts(value):
    return [part.strip() for part in value.split(",") if part.strip()]


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
        newline="\n",
    )


def export_bom(wb, generated_at):
    grouped = {}
    for parent, children in read_rows(wb["BOM"], 2):
        if not parent or not children:
            continue
        grouped.setdefault(parent, [])
        grouped[parent].extend(split_parts(children))

    bom = [
        {"parent_number": parent, "child_parts": ",".join(unique_list(grouped[parent]))}
        for parent in sorted(grouped)
    ]
    fixed = sorted({part for (part,) in read_rows(wb["固定料"], 1) if part})
    return {
        "generated_at": generated_at,
        "counts": {"bom": len(bom), "fixed": len(fixed)},
        "bom": bom,
        "fixed": fixed,
    }


def export_packaging(wb, generated_at):
    products = {}
    for product_number, specification, package_1 in read_rows(wb["包裝"], 3):
        if product_number:
            products[product_number] = {
                "product_number": product_number,
                "specification": specification,
                "package_1": package_1,
            }

    packaging = [products[key] for key in sorted(products)]
    return {
        "generated_at": generated_at,
        "count": len(packaging),
        "packaging": packaging,
    }


def export_usage(wb, output_dir, generated_at):
    usage_dir = output_dir / "usage"
    if usage_dir.exists():
        shutil.rmtree(usage_dir)
    usage_dir.mkdir(parents=True, exist_ok=True)

    groups = {}
    for parent_number, child_part_number, batch_quantity, item_name in read_rows(wb["用量"], 4):
        if not parent_number or not child_part_number:
            continue
        groups.setdefault(parent_number, []).append({
            "child_part_number": child_part_number,
            "batch_quantity": batch_quantity,
            "item_name": item_name,
        })

    manifest = {
        "generated_at": generated_at,
        "count": sum(len(rows) for rows in groups.values()),
        "parents": {},
    }

    for index, parent_number in enumerate(sorted(groups), start=1):
        filename = f"u{index:05d}.json"
        rows = sorted(groups[parent_number], key=lambda row: row["child_part_number"])
        write_json(usage_dir / filename, {
            "generated_at": generated_at,
            "parent_number": parent_number,
            "count": len(rows),
            "rows": rows,
        })
        manifest["parents"][parent_number] = filename

    write_json(usage_dir / "manifest.json", manifest)
    return {"parents": len(groups), "rows": manifest["count"]}


def main():
    parser = argparse.ArgumentParser(description="Export BOM.xlsx to static JSON files for Worker Assets.")
    parser.add_argument("--workbook", default="./文件/BOM.xlsx")
    parser.add_argument("--output", default="./public/data")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    output_dir = Path(args.output)
    generated_at = datetime.now(timezone.utc).isoformat()

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet in ["BOM", "固定料", "用量", "包裝"]:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"缺少工作表: {sheet}")

    output_dir.mkdir(parents=True, exist_ok=True)
    bom_payload = export_bom(wb, generated_at)
    packaging_payload = export_packaging(wb, generated_at)
    usage_counts = export_usage(wb, output_dir, generated_at)

    write_json(output_dir / "bom.json", bom_payload)
    write_json(output_dir / "packaging.json", packaging_payload)

    print(f"已輸出 {output_dir}")
    print(f"BOM: {bom_payload['counts']['bom']} 筆")
    print(f"固定料: {bom_payload['counts']['fixed']} 筆")
    print(f"用量: {usage_counts['rows']} 筆 / {usage_counts['parents']} 個母件")
    print(f"包裝: {packaging_payload['count']} 筆")


if __name__ == "__main__":
    main()
