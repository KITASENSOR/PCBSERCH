import argparse
from pathlib import Path

import openpyxl


SHEET_ALIASES = {
    "all": "all",
    "全部": "all",
    "bom": "BOM",
    "BOM": "BOM",
    "fixed": "固定料",
    "固定料": "固定料",
    "usage": "用量",
    "用量": "用量",
    "packaging": "包裝",
    "package": "包裝",
    "包裝": "包裝",
}


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def sql_text(value):
    if value == "":
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def read_rows(ws, columns):
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [cell_text(row[i]) if i < len(row) else "" for i in range(columns)]
        if any(values):
            yield values


def write_insert(f, table, fields, rows, batch_size=250):
    rows = list(rows)
    if not rows:
        return 0

    for start in range(0, len(rows), batch_size):
        chunk = rows[start:start + batch_size]
        f.write(f"INSERT INTO {table} ({', '.join(fields)}) VALUES\n")
        values = []
        for row in chunk:
            values.append("(" + ", ".join(sql_text(value) for value in row) + ")")
        f.write(",\n".join(values))
        f.write(";\n")

    return len(rows)


def unique_list(values):
    seen = set()
    result = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def split_parts(value):
    return [part.strip() for part in value.split(",") if part.strip()]


def normalize_sheet(value):
    key = str(value or "all").strip()
    if key not in SHEET_ALIASES:
        allowed = ", ".join(["all", "bom", "fixed", "usage", "packaging"])
        raise SystemExit(f"不支援的工作表參數: {value}。可用參數: {allowed}")
    return SHEET_ALIASES[key]


def ensure_sheets(wb, requested_sheet):
    required_sheets = ["BOM", "固定料", "用量", "包裝"] if requested_sheet == "all" else [requested_sheet]
    missing = [name for name in required_sheets if name not in wb.sheetnames]
    if missing:
        raise SystemExit(f"缺少工作表: {', '.join(missing)}")


def export_bom(f, wb):
    f.write("DELETE FROM bom;\n")
    grouped = {}
    for parent, children in read_rows(wb["BOM"], 2):
        if not parent or not children:
            continue
        grouped.setdefault(parent, [])
        grouped[parent].extend(split_parts(children))

    rows = []
    for parent in sorted(grouped):
        rows.append([parent, ",".join(unique_list(grouped[parent]))])

    return write_insert(
        f,
        "bom",
        ["parent_number", "child_parts"],
        rows,
    )


def export_fixed_parts(f, wb):
    f.write("DELETE FROM fixed_parts;\n")
    parts = sorted({part for (part,) in read_rows(wb["固定料"], 1) if part})
    return write_insert(
        f,
        "fixed_parts",
        ["part_number"],
        ([part] for part in parts),
    )


def export_usage(f, wb):
    f.write("DELETE FROM usage_items;\n")
    f.write("DELETE FROM sqlite_sequence WHERE name = 'usage_items';\n")
    return write_insert(
        f,
        "usage_items",
        ["parent_number", "child_part_number", "batch_quantity", "item_name"],
        read_rows(wb["用量"], 4),
    )


def export_packaging(f, wb):
    f.write("DELETE FROM packaging;\n")
    products = {}
    for product_number, specification, package_1 in read_rows(wb["包裝"], 3):
        if product_number:
            products[product_number] = [product_number, specification, package_1]

    return write_insert(
        f,
        "packaging",
        ["product_number", "specification", "package_1"],
        [products[key] for key in sorted(products)],
    )


def main():
    parser = argparse.ArgumentParser(description="Export BOM.xlsx sheets to a D1 seed SQL file.")
    parser.add_argument("--workbook", default="./BOM.xlsx")
    parser.add_argument("--output", default="./data/bom_seed.sql")
    parser.add_argument("--sheet", default="all", help="all, bom, fixed, usage, packaging")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    requested_sheet = normalize_sheet(args.sheet)
    ensure_sheets(wb, requested_sheet)

    with output_path.open("w", encoding="utf-8", newline="\n") as f:
        counts = {}

        if requested_sheet in ("all", "包裝"):
            counts["包裝"] = export_packaging(f, wb)
        if requested_sheet in ("all", "用量"):
            counts["用量"] = export_usage(f, wb)
        if requested_sheet in ("all", "固定料"):
            counts["固定料"] = export_fixed_parts(f, wb)
        if requested_sheet in ("all", "BOM"):
            counts["BOM"] = export_bom(f, wb)

    print(f"已輸出 {output_path}")
    for name, count in counts.items():
        print(f"{name}: {count} 筆")


if __name__ == "__main__":
    main()
